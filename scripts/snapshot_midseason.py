"""
Mid-season snapshot: scrape Cricbattle scores and export to Excel.
Authenticates with Cricbattle, fetches all available match scores,
aggregates by player, merges with player roster CSV, and writes Excel.
"""

import asyncio
import json
import sys
from decimal import Decimal
from pathlib import Path

import httpx
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EMAIL = "vinod492@gmail.com"
PASSWORD = "Cricket@2026"
LEAGUE_ID = "676423"

LOGIN_START = "https://www.cricbattle.com/Account/LoginRegister/ByEmail"
LOGIN_EMAIL = "https://www.cricbattle.com/Account/Login/ByEmail"
LOGIN_OPTIONS = "https://www.cricbattle.com/Account/Login/Options"
SCORES_PAGE = f"https://fantasycricket.cricbattle.com/MyFantasy/Player-Scores-Breakdown?LeagueModel=SalaryCap&LeagueId={LEAGUE_ID}"
SCORES_API = "https://fantasycricket.cricbattle.com/MyFantasy/PlayerScoresBreakdown/GetLeaguePlayerScoresBreakdownData"

PLAYERS_CSV = Path(__file__).parent.parent / "players" / "ipl_2026_players_latest.csv"
OUTPUT = Path(__file__).parent.parent / "players" / "midseason_snapshot_2026.xlsx"


async def login() -> httpx.AsyncClient:
    client = httpx.AsyncClient(
        follow_redirects=True,
        timeout=httpx.Timeout(30.0),
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"},
    )
    await client.get(LOGIN_START)
    resp = await client.post(LOGIN_EMAIL, data={"IsUseOfficialEmail": "False", "CountryId": "3", "Email": EMAIL})
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")
    form = soup.find("form", {"action": "/Account/Login/Options"})
    if not form:
        raise RuntimeError(f"Login form not found (landed on: {resp.url})")
    hidden = {inp.get("name"): inp.get("value", "") for inp in form.find_all("input") if inp.get("name")}
    hidden["Password"] = PASSWORD
    hidden["IsPassword"] = "True"

    login_resp = await client.post(LOGIN_OPTIONS, data=hidden)
    login_resp.raise_for_status()
    if ".CBProd" not in client.cookies:
        raise RuntimeError("Login failed: auth cookie not set")

    print("✓ Logged in to Cricbattle")
    return client


async def get_matches(client: httpx.AsyncClient) -> list[dict]:
    resp = await client.get(SCORES_PAGE)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    select = soup.find("select", {"id": "MatchId"})
    matches = []
    if select:
        for opt in select.find_all("option"):
            val = opt.get("value", "").strip()
            label = opt.get_text(strip=True)
            if val and val not in ("0", ""):
                matches.append({"match_id": val, "match_label": label})
    print(f"✓ Found {len(matches)} match(es): {[m['match_label'] for m in matches]}")
    return matches


async def get_match_scores(client: httpx.AsyncClient, match_id: str, match_label: str) -> list[dict]:
    resp = await client.post(
        SCORES_API,
        content=json.dumps({"lid": LEAGUE_ID, "matchid": match_id}),
        headers={"Content-Type": "application/json; charset=utf-8"},
    )
    resp.raise_for_status()
    data = resp.json()
    raw = (data.get("Result") or {}).get("lstPlayer") or []
    rows = []
    for p in raw:
        name = (p.get("PlayerName") or "").strip()
        if not name:
            continue
        pts = Decimal(str(p.get("TotalScore") or 0))
        fours = sum(inning.get("Fours", 0) or 0 for inning in p.get("lstScore", []))
        sixes = sum(inning.get("Sixes", 0) or 0 for inning in p.get("lstScore", []))
        rows.append({"player_name": name, "points": float(pts), "fours": fours, "sixes": sixes, "match": match_label})
    print(f"  ✓ {match_label}: {len(rows)} players scored")
    return rows


def build_excel(agg: pd.DataFrame, match_labels: list[str], per_match: pd.DataFrame):
    wb = Workbook()

    # ── Sheet 1: Rankings ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Mid-Season Rankings"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    rank_fill = PatternFill("solid", start_color="FFD700")

    headers = ["Rank", "Player Name", "IPL Team", "Designation", "Total Points", "Fours", "Sixes", "Boundary Pts"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    col_widths = [6, 26, 28, 24, 14, 8, 8, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, row in enumerate(agg.itertuples(), start=1):
        r = idx + 1
        fill = alt_fill if idx % 2 == 0 else None
        base_font = Font(name="Arial", size=10)

        rank_cell = ws.cell(row=r, column=1, value=idx)
        rank_cell.font = Font(name="Arial", size=10, bold=True)
        rank_cell.alignment = Alignment(horizontal="center")
        if idx <= 3:
            rank_cell.fill = rank_fill

        for col, val in enumerate([row.player_name, row.ipl_team, row.designation,
                                    row.points, row.fours, row.sixes], start=2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = base_font
            if fill:
                cell.fill = fill
            if col in (5, 6, 7):
                cell.alignment = Alignment(horizontal="center")

        # Boundary Pts: 0.5*fours + 2*sixes
        bp_val = round(row.fours * 0.5 + row.sixes * 2, 2)
        bp_cell = ws.cell(row=r, column=8, value=bp_val)
        bp_cell.font = base_font
        bp_cell.alignment = Alignment(horizontal="center")
        if fill:
            bp_cell.fill = fill

    ws.freeze_panes = "A2"

    # ── Sheet 2: Per-Match Breakdown ───────────────────────────────────────
    ws2 = wb.create_sheet("Per-Match Breakdown")
    pm_headers = ["Player Name", "IPL Team", "Designation"] + match_labels + ["Total Points"]
    for col, h in enumerate(pm_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 24
    for i in range(4, len(pm_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    for idx, row in enumerate(per_match.itertuples(), start=2):
        fill = alt_fill if (idx - 1) % 2 == 0 else None
        ws2.cell(row=idx, column=1, value=row.player_name).font = Font(name="Arial", size=10)
        ws2.cell(row=idx, column=2, value=row.ipl_team).font = Font(name="Arial", size=10)
        ws2.cell(row=idx, column=3, value=row.designation).font = Font(name="Arial", size=10)
        for m_idx, label in enumerate(match_labels, start=4):
            val = getattr(row, f"pts_{m_idx - 4}", 0)
            cell = ws2.cell(row=idx, column=m_idx, value=val or 0)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

        match_pts = [getattr(row, f"pts_{i}", 0) or 0 for i in range(len(match_labels))]
        total_cell = ws2.cell(row=idx, column=4 + len(match_labels),
                               value=round(sum(match_pts), 2))
        total_cell.font = Font(name="Arial", size=10, bold=True)
        total_cell.alignment = Alignment(horizontal="center")
        if fill:
            total_cell.fill = fill

    ws2.freeze_panes = "D2"

    wb.save(OUTPUT)
    print(f"\n✓ Saved: {OUTPUT}")


async def main():
    client = await login()

    matches = await get_matches(client)
    if not matches:
        print("No matches found in Cricbattle league. Scores may not be available yet.")
        sys.exit(1)

    # Fetch all match scores
    all_rows = []
    for m in matches:
        rows = await get_match_scores(client, m["match_id"], m["match_label"])
        for row in rows:
            row["match_idx"] = matches.index(m)
        all_rows.extend(rows)

    await client.aclose()

    if not all_rows:
        print("No score data returned. Exiting.")
        sys.exit(1)

    # Load roster
    roster = pd.read_csv(PLAYERS_CSV)
    roster.columns = [c.strip() for c in roster.columns]
    roster = roster.rename(columns={"Player Name": "player_name", "Team": "ipl_team", "Designation": "designation", "Ranking": "ranking"})
    roster["player_name_lower"] = roster["player_name"].str.lower().str.strip()

    # Aggregate scores across all matches
    scores_df = pd.DataFrame(all_rows)
    agg = scores_df.groupby("player_name").agg(points=("points", "sum"), fours=("fours", "sum"), sixes=("sixes", "sum")).reset_index()

    # Merge with roster
    agg["player_name_lower"] = agg["player_name"].str.lower().str.strip()
    merged = agg.merge(roster[["player_name_lower", "player_name", "ipl_team", "designation", "ranking"]],
                       on="player_name_lower", how="left", suffixes=("_cricbattle", ""))

    # Use cricbattle name as fallback
    merged["player_name"] = merged["player_name"].fillna(merged["player_name_cricbattle"])
    merged = merged.sort_values("points", ascending=False).reset_index(drop=True)
    merged = merged[["player_name", "ipl_team", "designation", "points", "fours", "sixes"]].fillna({"ipl_team": "Unknown", "designation": "Unknown"})

    # Build per-match breakdown
    match_labels = [m["match_label"] for m in matches]
    pivot = scores_df.pivot_table(index="player_name", columns="match_idx", values="points", aggfunc="sum", fill_value=0).reset_index()
    pivot.columns = ["player_name"] + [f"pts_{i}" for i in range(len(matches))]
    per_match = merged.merge(pivot, on="player_name", how="left")

    print(f"\nTop 10 players by points:")
    print(merged[["player_name", "ipl_team", "points"]].head(10).to_string(index=False))

    build_excel(merged, match_labels, per_match)


if __name__ == "__main__":
    asyncio.run(main())
